"""
Excel export for a single calculation (spec §5 action bar / §10 task 7).

Produces a styled .xlsx in-memory (bytes) so it can be served via
st.download_button without touching disk.
"""

from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from spot_risk_core import SpotRiskResult

_NAVY = "0F1B33"
_ACCENT = "2563EB"
_GREEN = "16A34A"
_RED = "DC2626"
_LIGHT = "F1F5F9"


def build_excel(*, ticker: str, mode: str, setup_date, capital: float,
                risk_pct: float, support: float, resistance: float,
                deviation_pct: float, atr: float, result: SpotRiskResult) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Spot Risk Plan"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22

    thin = Side(style="thin", color="D7DEE8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header(row, text):
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, color="FFFFFF", size=12)
        c.fill = PatternFill("solid", fgColor=_NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c2 = ws.cell(row=row, column=2, value="")
        c2.fill = PatternFill("solid", fgColor=_NAVY)
        ws.row_dimensions[row].height = 22

    def kv(row, label, value, *, color=None, bold=False, fmt=None):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(color="334155")
        lc.border = border
        lc.fill = PatternFill("solid", fgColor=_LIGHT)
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = Font(bold=bold, color=color or "0F172A", size=11)
        vc.border = border
        vc.alignment = Alignment(horizontal="right")
        if fmt:
            vc.number_format = fmt

    # Title
    t = ws.cell(row=1, column=1, value=f"Spot Risk Game Plan — {ticker} · {mode}")
    t.font = Font(bold=True, size=15, color=_ACCENT)
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 28

    sd = setup_date.isoformat() if isinstance(setup_date, date) else str(setup_date)

    r = 3
    header(r, "Setup"); r += 1
    for lbl, val, fmt in [
        ("Ticker", ticker, None), ("Mode", mode, None), ("Setup Date", sd, None),
        ("Capital (USD)", capital, "#,##0.00"),
        ("Risk %", risk_pct / 100, "0.00%"),
        ("Deviation %", deviation_pct / 100, "0.00%"),
        ("Support", support, "#,##0.00"),
        ("Resistance", resistance, "#,##0.00"),
        ("ATR(10)", atr, "#,##0.00"),
    ]:
        kv(r, lbl, val, fmt=fmt); r += 1

    r += 1
    header(r, "Plan"); r += 1
    kv(r, "Entry Price", result.entry, bold=True, color=_ACCENT, fmt="#,##0.00"); r += 1
    kv(r, "Cut Loss", result.cut_loss, bold=True, color=_RED, fmt="#,##0.00"); r += 1
    kv(r, "Take Profit", result.take_profit, bold=True, color=_GREEN, fmt="#,##0.00"); r += 1
    kv(r, "Risk / share", result.risk_ps, fmt="#,##0.00"); r += 1
    kv(r, "Reward / share", result.reward_ps, fmt="#,##0.00"); r += 1
    kv(r, "Reward : Risk", result.rr, bold=True, fmt="0.00"); r += 1

    r += 1
    header(r, "Sizing"); r += 1
    kv(r, "Quantity (shares)", result.qty, bold=True, fmt="#,##0"); r += 1
    kv(r, "Capital Required", result.capital_req, fmt="#,##0.00"); r += 1
    kv(r, "Total Risk", result.total_risk, color=_RED, fmt="#,##0.00"); r += 1
    kv(r, "Total Win", result.total_win, color=_GREEN, fmt="#,##0.00"); r += 1
    kv(r, "% Risk of Account", result.risk_acct_pct / 100, fmt="0.00%"); r += 1
    kv(r, "% Win of Account", result.win_acct_pct / 100, fmt="0.00%"); r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
